from __future__ import annotations

import argparse
import math
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


MONTHS = {
    "januar": 1, "februar": 2, "märz": 3, "maerz": 3, "mrz": 3,
    "april": 4, "mai": 5, "juni": 6, "juli": 7, "august": 8,
    "september": 9, "oktober": 10, "november": 11, "dezember": 12,
}
MONTH_LABELS = ["Jan.", "Feb.", "März", "April", "Mai", "Juni", "Juli", "Aug.", "Sept.", "Okt.", "Nov.", "Dez."]


def normalized(value: object) -> str:
    return str(value or "").strip().rstrip(":").casefold()


def month_headers(ws) -> list[tuple[int, int]]:
    result = []
    for row in range(1, ws.max_row + 1):
        month = MONTHS.get(normalized(ws.cell(row, 1).value))
        if month:
            result.append((row, month))
    return result


def total_rows_by_month(ws) -> dict[int, list[int]]:
    headers = month_headers(ws)
    result = {month: [] for _, month in headers}
    for index, (start, month) in enumerate(headers):
        end = headers[index + 1][0] - 1 if index + 1 < len(headers) else ws.max_row
        for row in range(start + 1, end + 1):
            if normalized(ws.cell(row, 1).value) == "gesamt" and isinstance(ws.cell(row, 9).value, (int, float)):
                result[month].append(row)
    return result


def inferred_count(ws_values, ws_formulas, row: int) -> int:
    total = float(ws_values.cell(row, 9).value or 0)
    average = ws_values.cell(row, 13).value
    if isinstance(average, (int, float)) and average > 0:
        return max(1, round(total / float(average)))

    formula = str(ws_formulas.cell(row, 9).value or "")
    match = re.fullmatch(r"=SUM\(I(\d+):I(\d+)\)", formula, flags=re.IGNORECASE)
    if not match:
        return 0
    start, end = map(int, match.groups())
    excluded = {"prämie", "praemie", "übernachtungskosten", "uebernachtungskosten"}
    count = 0
    for source_row in range(start, end + 1):
        label = normalized(ws_values.cell(source_row, 1).value).replace("ä", "ae").replace("ü", "ue")
        amount = ws_values.cell(source_row, 9).value
        if label not in excluded and isinstance(amount, (int, float)) and amount > 0:
            count += 1
    return count


def period_summary(wb_values, wb_formulas, year: int, through_month: int) -> dict[str, float | int | str]:
    ws_values = wb_values[str(year)]
    ws_formulas = wb_formulas[str(year)]
    grouped = total_rows_by_month(ws_values)
    rows = [row for month in range(1, through_month + 1) for row in grouped.get(month, [])]
    total = sum(float(ws_values.cell(row, 9).value or 0) for row in rows)
    count = sum(inferred_count(ws_values, ws_formulas, row) for row in rows)
    label = f"Jan.–{MONTH_LABELS[through_month - 1]} {year}"
    return {
        "year": year,
        "period": label,
        "months": through_month,
        "ytd_net": round(total, 2),
        "annualized_net": round(total / through_month * 12, 2),
        "average_net": round(total / count, 2) if count else 0,
    }


def latest_completed_month(ws) -> int:
    grouped = total_rows_by_month(ws)
    completed = [month for month, rows in grouped.items() if rows]
    if not completed:
        raise RuntimeError("Im aktuellen Jahresblatt wurde keine blaue Gesamtzeile gefunden.")
    return max(completed)


def private_totals(wb_values, years: tuple[int, int]) -> dict[int, float]:
    sheet_name = next(name for name in wb_values.sheetnames if normalized(name).startswith("privatauftr"))
    ws = wb_values[sheet_name]
    totals = {year: 0.0 for year in years}
    active_year = None
    for row in range(1, ws.max_row + 1):
        value_a = ws.cell(row, 1).value
        if isinstance(value_a, (int, float)) and int(value_a) in totals:
            active_year = int(value_a)
            continue
        if active_year is None or normalized(value_a) == "gesamt":
            continue
        amount = ws.cell(row, 4).value
        if isinstance(amount, (int, float)):
            totals[active_year] += float(amount)
    return {year: round(value, 2) for year, value in totals.items()}


def php_number(value: float) -> str:
    return f"{value:.2f}"


def write_endpoint(output: Path, source: Path, current: dict, previous: dict, private: dict[int, float]) -> None:
    stamp = datetime.fromtimestamp(source.stat().st_mtime).strftime("%d.%m.%Y %H:%M")
    content = f'''<?php
declare(strict_types=1);

return [
    'source' => '{source.name}',
    'source_updated_at' => '{stamp}',
    'comparison' => 'gleicher Zeitraum',
    'current' => [
        'year' => {current['year']},
        'period' => '{current['period']}',
        'months' => {current['months']},
        'ytd_net' => {php_number(current['ytd_net'])},
        'annualized_net' => {php_number(current['annualized_net'])},
        'average_net' => {php_number(current['average_net'])},
        'private_gross' => {php_number(private[current['year']])},
    ],
    'previous' => [
        'year' => {previous['year']},
        'period' => '{previous['period']}',
        'months' => {previous['months']},
        'ytd_net' => {php_number(previous['ytd_net'])},
        'annualized_net' => {php_number(previous['annualized_net'])},
        'average_net' => {php_number(previous['average_net'])},
        'private_gross' => {php_number(private[previous['year']])},
    ],
];
'''
    output.write_text(content, encoding="utf-8", newline="\n")


def main() -> None:
    parser = argparse.ArgumentParser(description="Aktualisiert Christians Portal-Umsatzübersicht aus der Claims-Arbeitsmappe.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output", type=Path, default=Path("public/intern/api/revenue-summary-fallback.php"))
    args = parser.parse_args()
    workbook = args.workbook.resolve()
    if not workbook.is_file():
        raise FileNotFoundError(workbook)

    current_year = datetime.now().year
    previous_year = current_year - 1
    wb_values = load_workbook(workbook, data_only=True)
    wb_formulas = load_workbook(workbook, data_only=False)
    through_month = latest_completed_month(wb_values[str(current_year)])
    current = period_summary(wb_values, wb_formulas, current_year, through_month)
    previous = period_summary(wb_values, wb_formulas, previous_year, through_month)
    private = private_totals(wb_values, (current_year, previous_year))
    write_endpoint(args.output.resolve(), workbook, current, previous, private)
    print(f"{current['period']}: {current['ytd_net']:.2f} EUR netto, Ø {current['average_net']:.2f} EUR, Jahreswert {current['annualized_net']:.2f} EUR")
    print(f"{previous['period']}: {previous['ytd_net']:.2f} EUR netto, Ø {previous['average_net']:.2f} EUR, Jahreswert {previous['annualized_net']:.2f} EUR")
    print(f"Privataufträge {current_year}: {private[current_year]:.2f} EUR brutto; {previous_year}: {private[previous_year]:.2f} EUR brutto")


if __name__ == "__main__":
    main()
